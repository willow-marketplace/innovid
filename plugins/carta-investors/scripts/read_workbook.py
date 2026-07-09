# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
Read an .xlsx file and emit a JSON description of its sheets, used by
the budget skills when they run outside Claude for Excel (Claude Code /
Cowork / any local-file context).

Usage:
    uv run read_workbook.py <path> [--sheet <name>] [--max-rows N]

Output (stdout): JSON. Claude consumes this to figure out section
boundaries, header rows, line items, and any existing actuals / budget
columns. The script stays intentionally dumb — emits raw cell content,
no inference.
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def describe_workbook(path: Path, sheet_filter: str | None, max_rows: int) -> dict:
    wb = load_workbook(filename=path, read_only=True, data_only=False)
    sheets = []
    for name in wb.sheetnames:
        if sheet_filter and name != sheet_filter:
            continue
        ws = wb[name]
        rows = []
        for r, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if r > max_rows:
                break
            cells = {}
            for cell in row:
                if cell.value is None:
                    continue
                value = cell.value
                is_formula = isinstance(value, str) and value.startswith("=")
                cells[cell.coordinate] = {
                    "value": str(value),
                    "is_formula": is_formula,
                    "number_format": cell.number_format,
                }
            if cells:
                rows.append({"row": r, "cells": cells})
        sheets.append({
            "name": name,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "sampled_rows": rows,
        })
    return {"workbook_path": str(path), "sheets": sheets}


def main(argv):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", type=Path, help="Path to the .xlsx file")
    parser.add_argument("--sheet", help="Limit output to a single sheet")
    parser.add_argument("--max-rows", type=int, default=300,
                        help="Cap per-sheet row sample (default 300)")
    args = parser.parse_args(argv)
    if not args.path.exists():
        print(f"error: file not found: {args.path}", file=sys.stderr)
        return 1
    result = describe_workbook(args.path, args.sheet, args.max_rows)
    print(json.dumps(result, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
