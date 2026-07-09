# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1", "Pillow>=10.0"]
# ///
"""
Apply a JSON payload of operations to an .xlsx file (creating it if
missing). Used by the budget skills when they run outside Claude for
Excel (Claude Code / Cowork / any local-file context).

Usage:
    uv run write_workbook.py <payload_json_path>
    uv run write_workbook.py --stdin   # read payload from stdin

Operations supported:
    create_sheet                — create a new tab
    delete_sheet                — remove an existing tab
    write_cell                  — single-cell value
    write_formula               — single-cell formula
    write_range                 — 2D values from a starting cell
    fill_formula_horizontal     — seed formula at start, translate across a row
    fill_formula_vertical       — seed formula at start, translate down a column
    set_format                  — number format on a range
    set_bold                    — bold font on a range
    freeze_panes                — freeze at a ref (use sparingly; not for carta-create-budget)
    set_column_width            — fixed width (use sparingly; prefer autofit_columns)
    autofit_columns             — compute width per column from cell content
    add_image                   — anchor a PNG at a cell, sized to N rows tall
    set_comment                 — attach a native cell comment (used for sparse-history flagging)
    merge_cells                 — merge a cell range (used for period header rows in tag-view layout)

Payload shape:
    {
      "workbook_path": "/path/to/file.xlsx",
      "create_if_missing": true,
      "operations": [
        {"op": "create_sheet", "name": "Budget FY2026", "position": 0},
        {"op": "write_cell", "sheet": "Budget FY2026", "ref": "A1",
         "value": "Example Capital, LLC"},
        {"op": "write_cell", "sheet": "Budget FY2026", "ref": "A2",
         "value": "2026 Budget"},
        {"op": "write_range", "sheet": "Budget FY2026", "start": "A6",
         "values": [["Account", "Jan 2026 B", "Jan 2026 A", "Jan 2026 V", "..."]]},
        {"op": "fill_formula_horizontal", "sheet": "Budget FY2026",
         "start": "D7", "end": "AN7", "formula": "=B7-C7"},
        {"op": "fill_formula_vertical", "sheet": "Budget FY2026",
         "start": "N7", "end": "N50", "formula": "=SUM(B7:M7)"},
        {"op": "set_format", "sheet": "Budget FY2026", "ref": "B7:N50",
         "number_format": "_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* \"-\"??_);_(@_)"},
        {"op": "set_bold", "sheet": "Budget FY2026", "ref": "A6:N6"},
        {"op": "autofit_columns", "sheet": "Budget FY2026", "columns": "B:AN"}
      ]
    }

Notes for the budget skills:
    - Use the [$$-en-US] locale token in number_format, NOT a bare "$".
      Bare-$ renders in system locale (R$ on pt-BR, etc.).
    - Do NOT include a freeze_panes op for carta-create-budget output — the
      Carta standard (see carta-consolidating-pnl/references/formatting.md)
      does not freeze.
    - Prefer autofit_columns to set_column_width for accounting-format
      currency cells — fixed widths < 16pt show ##### for 5+ digit values.
    - fill_formula_horizontal / fill_formula_vertical use openpyxl's
      Translator to shift relative references, so a seed formula
      "=B7-C7" at D7 becomes "=E7-F7" at G7 (shifted right by 3 cols).
    - add_image anchors a PNG at the given cell and sizes its height to
      `rows` default-height rows (15pt each). Width is derived from the
      PNG's native aspect ratio so the logo stays proportional. Used by
      carta-create-budget to brand both tabs with "Powered by Carta".
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font
from openpyxl.utils import (
    column_index_from_string,
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from PIL import Image as PILImage

# Default row height in openpyxl is 15 points (Excel's default for Calibri 11).
# A point in EMUs is 12700; openpyxl Image width/height are in pixels at 96 DPI,
# so 1 pt == 4/3 px. We do the conversion here so callers can pass row counts.
_DEFAULT_ROW_HEIGHT_PT = 15.0
_PT_TO_PX = 4.0 / 3.0


def _autofit_width_for_cell(cell) -> int:
    """Approximate display width for one cell — string length + small padding
    for accounting-format currency."""
    if cell.value is None:
        return 0
    val = str(cell.value)
    if val.startswith("="):
        # Formula — we can't know the rendered length without recalc.
        # Estimate ~14 chars (enough for currency through $billions).
        return 14
    n = len(val)
    fmt = cell.number_format or ""
    # Accounting / currency formats add parens, $ symbol, separators
    if "$" in fmt or "(#,##0" in fmt:
        n = max(n, 12)
    return n


def apply_operations(payload: dict) -> dict:
    path = Path(payload["workbook_path"])
    create_if_missing = payload.get("create_if_missing", True)

    if path.exists():
        wb = load_workbook(filename=path)
    elif create_if_missing:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]
    else:
        raise FileNotFoundError(f"workbook not found: {path}")

    results = []
    for op in payload.get("operations", []):
        kind = op["op"]
        if kind == "create_sheet":
            name = op["name"]
            position = op.get("position")
            if name in wb.sheetnames:
                results.append({"op": kind, "name": name, "status": "exists"})
                continue
            wb.create_sheet(title=name, index=position)
            results.append({"op": kind, "name": name, "status": "created"})
        elif kind == "delete_sheet":
            name = op["sheet"]
            if name in wb.sheetnames:
                del wb[name]
                results.append({"op": kind, "sheet": name, "status": "deleted"})
            else:
                results.append({"op": kind, "sheet": name, "status": "not_found"})
        elif kind == "write_cell":
            ws = wb[op["sheet"]]
            ws[op["ref"]] = op["value"]
            results.append({"op": kind, "ref": op["ref"], "status": "ok"})
        elif kind == "write_formula":
            ws = wb[op["sheet"]]
            ws[op["ref"]] = op["formula"]
            results.append({"op": kind, "ref": op["ref"], "status": "ok"})
        elif kind == "write_range":
            ws = wb[op["sheet"]]
            row0, col0 = coordinate_to_tuple(op["start"])
            for i, row in enumerate(op["values"]):
                for j, value in enumerate(row):
                    ws.cell(row=row0 + i, column=col0 + j, value=value)
            results.append({"op": kind, "start": op["start"],
                            "rows": len(op["values"]), "status": "ok"})
        elif kind == "fill_formula_horizontal":
            ws = wb[op["sheet"]]
            start = op["start"]
            end = op["end"]
            formula = op["formula"]
            start_row, start_col = coordinate_to_tuple(start)
            end_row, end_col = coordinate_to_tuple(end)
            if start_row != end_row:
                raise ValueError(
                    f"fill_formula_horizontal: start and end must be on the "
                    f"same row ({start} → {end})"
                )
            ws[start] = formula
            for col in range(start_col + 1, end_col + 1):
                target = f"{get_column_letter(col)}{start_row}"
                ws[target] = Translator(formula, origin=start).translate_formula(target)
            results.append({
                "op": kind, "start": start, "end": end,
                "cells": end_col - start_col + 1, "status": "ok",
            })
        elif kind == "fill_formula_vertical":
            ws = wb[op["sheet"]]
            start = op["start"]
            end = op["end"]
            formula = op["formula"]
            start_row, start_col = coordinate_to_tuple(start)
            end_row, end_col = coordinate_to_tuple(end)
            if start_col != end_col:
                raise ValueError(
                    f"fill_formula_vertical: start and end must be in the "
                    f"same column ({start} → {end})"
                )
            ws[start] = formula
            for row in range(start_row + 1, end_row + 1):
                target = f"{get_column_letter(start_col)}{row}"
                ws[target] = Translator(formula, origin=start).translate_formula(target)
            results.append({
                "op": kind, "start": start, "end": end,
                "cells": end_row - start_row + 1, "status": "ok",
            })
        elif kind == "set_format":
            ws = wb[op["sheet"]]
            for row in ws[op["ref"]]:
                for cell in row:
                    cell.number_format = op["number_format"]
            results.append({"op": kind, "ref": op["ref"], "status": "ok"})
        elif kind == "set_bold":
            ws = wb[op["sheet"]]
            for row in ws[op["ref"]]:
                for cell in row:
                    cell.font = Font(bold=True)
            results.append({"op": kind, "ref": op["ref"], "status": "ok"})
        elif kind == "freeze_panes":
            ws = wb[op["sheet"]]
            ws.freeze_panes = op["ref"]
            results.append({"op": kind, "ref": op["ref"], "status": "ok"})
        elif kind == "set_column_width":
            ws = wb[op["sheet"]]
            ws.column_dimensions[op["column"]].width = op["width"]
            results.append({"op": kind, "column": op["column"], "status": "ok"})
        elif kind == "autofit_columns":
            ws = wb[op["sheet"]]
            cols = op.get("columns")  # e.g. "B:AN" or None for all
            if cols:
                start_letter, end_letter = cols.split(":")
                start_idx = column_index_from_string(start_letter)
                end_idx = column_index_from_string(end_letter)
            else:
                start_idx = 1
                end_idx = ws.max_column
            min_width = op.get("min_width", 8)
            max_width = op.get("max_width", 50)
            padding = op.get("padding", 2)
            for col_idx in range(start_idx, end_idx + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    n = _autofit_width_for_cell(cell)
                    if n > max_len:
                        max_len = n
                width = min(max(max_len + padding, min_width), max_width)
                ws.column_dimensions[col_letter].width = width
            results.append({"op": kind, "cols": cols or "all", "status": "ok"})
        elif kind == "add_image":
            ws = wb[op["sheet"]]
            image_path = Path(op["path"])
            if not image_path.is_absolute():
                image_path = (path.parent / image_path).resolve()
            if not image_path.exists():
                results.append({"op": kind, "path": str(image_path),
                                "status": "missing"})
                continue
            anchor = op.get("anchor", "C1")
            rows = float(op.get("rows", 3))
            with PILImage.open(image_path) as native:
                native_w, native_h = native.size
            ratio = native_w / native_h if native_h else 1.0
            height_px = rows * _DEFAULT_ROW_HEIGHT_PT * _PT_TO_PX
            width_px = height_px * ratio
            img = OpenpyxlImage(str(image_path))
            img.height = height_px
            img.width = width_px
            img.anchor = anchor
            ws.add_image(img)
            results.append({"op": kind, "sheet": op["sheet"], "anchor": anchor,
                            "width": round(width_px, 2),
                            "height": round(height_px, 2),
                            "status": "ok"})
        elif kind == "set_comment":
            ws = wb[op["sheet"]]
            ws[op["ref"]].comment = Comment(op["text"], op.get("author", "Carta"))
            results.append({"op": kind, "sheet": op["sheet"], "ref": op["ref"],
                            "status": "ok"})
        elif kind == "merge_cells":
            ws = wb[op["sheet"]]
            ref = op["ref"]
            # Pre-check: openpyxl raises ValueError on overlapping ranges and
            # halts the whole apply_operations loop — wb.save() never runs and
            # the caller gets a terse stderr message with no `results` array to
            # diagnose which op failed. Unmerge any overlapping ranges first
            # (the new write will populate the merged region) and wrap the
            # merge call so a still-failing case returns a structured error
            # the model can see and retry against (e.g. by adding a
            # delete_sheet preamble on re-runs).
            try:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                for existing in list(ws.merged_cells.ranges):
                    overlaps = not (
                        max_col < existing.min_col
                        or min_col > existing.max_col
                        or max_row < existing.min_row
                        or min_row > existing.max_row
                    )
                    if overlaps:
                        ws.unmerge_cells(str(existing))
                ws.merge_cells(ref)
                results.append({"op": kind, "sheet": op["sheet"], "ref": ref,
                                "status": "ok"})
            except Exception as exc:
                results.append({"op": kind, "sheet": op["sheet"], "ref": ref,
                                "status": "error", "error": str(exc)})
        else:
            results.append({"op": kind, "status": "unknown_op"})

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return {
        "workbook_path": str(path),
        "operations_applied": len(results),
        "results": results,
    }


def main(argv):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("payload_path", nargs="?", type=Path,
                        help="Path to a JSON file with the operations payload")
    parser.add_argument("--stdin", action="store_true",
                        help="Read JSON payload from stdin instead of a file")
    args = parser.parse_args(argv)

    if args.stdin:
        payload = json.load(sys.stdin)
    elif args.payload_path:
        payload = json.loads(args.payload_path.read_text())
    else:
        print("error: provide payload path or --stdin", file=sys.stderr)
        return 1

    try:
        result = apply_operations(payload)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
