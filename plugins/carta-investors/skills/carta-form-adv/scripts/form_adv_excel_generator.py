# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
Form ADV Part 1A — Excel Filing Reference Generator
Reads form_adv_data.json, writes a multi-sheet .xlsx workbook.

Blue cells  = Carta-sourced value. Safe to use directly in IARD.
Orange cells = Must be entered manually in the IARD web portal.
"""

import json, argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ───────────────────────────────────────────────────────────
CARTA_FILL   = PatternFill("solid", fgColor="DBEAFE")   # light blue  — Carta-sourced
MANUAL_FILL  = PatternFill("solid", fgColor="FED7AA")   # light orange — manual IARD entry
HDR_FILL     = PatternFill("solid", fgColor="1E40AF")   # dark blue   — column headers
SECT_FILL    = PatternFill("solid", fgColor="E2E8F0")   # light grey  — section dividers
FUND_FILL    = PatternFill("solid", fgColor="1E40AF")   # dark blue   — fund name banner

CARTA_FONT   = Font(color="1E3A5F")
MANUAL_FONT  = Font(color="7C2D12", italic=True)
HDR_FONT     = Font(color="FFFFFF", bold=True)
BOLD         = Font(bold=True)
ITALIC       = Font(italic=True, color="6B7280")
THIN         = Border(bottom=Side(style="thin", color="CBD5E1"))


# ── Helpers ──────────────────────────────────────────────────────────────────

def fmt_currency(v):
    if v is None: return "—"
    try: return f"${float(v):,.0f}"
    except: return "—"

def fmt_pct(v, decimals=1):
    if v is None: return "—"
    try: return f"{float(v):.{decimals}f}%"
    except: return "—"

def fmt_multiple(v):
    if v is None: return "—"
    try: return f"{float(v):.2f}x"
    except: return "—"

def fmt_val(v):
    return "—" if (v is None or v == "") else str(v)

def carta(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = CARTA_FILL
    c.font = CARTA_FONT
    c.border = THIN
    return c

def manual(ws, row, col, label="[Enter in IARD]"):
    c = ws.cell(row=row, column=col, value=label)
    c.fill = MANUAL_FILL
    c.font = MANUAL_FONT
    c.border = THIN
    return c

def hdr(ws, row, values, col_widths=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center")
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def sect(ws, row, title, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = SECT_FILL
    c.font = Font(bold=True, color="1E293B")

def kv(ws, row, label, value, is_carta=True, val_col=2):
    ws.cell(row=row, column=1, value=label).font = BOLD
    if is_carta:
        carta(ws, row, val_col, value)
    else:
        manual(ws, row, val_col)


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_legend(wb, firm_name, reporting_date):
    ws = wb.create_sheet("Legend")
    hdr(ws, 1, [""], col_widths=[6, 30, 60])

    ws.cell(row=1, column=1, value=f"{firm_name} — Form ADV Part 1A Filing Reference").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Reporting date: {reporting_date}").font = ITALIC

    ws.cell(row=4, column=1, value="Colour key").font = BOLD
    for row, fill, font, label, desc in [
        (5, CARTA_FILL, CARTA_FONT,  "Carta-sourced",     "Value computed from Carta. Safe to enter directly in IARD."),
        (6, MANUAL_FILL, MANUAL_FONT, "Manual IARD entry", "Not available from Carta. Must be entered in the IARD web portal."),
    ]:
        c = ws.cell(row=row, column=2, value=label)
        c.fill = fill; c.font = font
        ws.cell(row=row, column=3, value=desc)

    ws.cell(row=8, column=1, value="Sheets").font = BOLD
    for i, (name, desc) in enumerate([
        ("Firm Overview",         "Items 5.D, 5.F, 5.H — regulatory AUM and non-US client %"),
        ("Per-Fund Detail",       "Schedule D §7.B.(1) — balance sheet, NAV, capital activity, investor counts per fund"),
        ("Investor Demographics", "US vs. non-US breakdown and entity type breakdown per fund"),
        ("Asset Composition",     "Schedule D §5.K.(1) — FMV by asset class"),
        ("Manual Fields",         "All fields requiring manual IARD entry, organized by ADV item"),
    ], 9):
        ws.cell(row=i, column=2, value=name).font = BOLD
        ws.cell(row=i, column=3, value=desc)

    ws.cell(row=15, column=1, value="Data caveats").font = Font(bold=True, color="B45309")
    for i, text in enumerate([
        "Non-US % (Item 5.H) is approximate — based on partner_country field in Carta. Verify completeness before filing.",
        "Investor counts: 'point-in-time' = as of reporting date. 'Snapshot' = current state; verify against subscription register.",
        "Legal fund names may differ from display names in Carta. Always verify against fund formation documents.",
        "Fund type classification is derived from investment_strategy_code — confirm legal classification with counsel.",
        "Regulatory AUM = Gross Assets + Unfunded Commitments (balance sheet method, per SEC instructions for private fund advisers).",
    ], 16):
        ws.cell(row=i, column=2, value=f"• {text}")


def _firm_lp_aggregates(firm_aggregates, demos):
    """Prefer SQL-computed firm_aggregates (distinct LPs). Fall back to summing per-fund
    counts if absent, which double-counts any LP committed to multiple funds — flag the
    fallback so the caller can surface a warning in the sheet.
    """
    fa = firm_aggregates or {}
    if fa and (fa.get("total_lp_investors") is not None or fa.get("us_lp_investors") is not None):
        return {
            "source": "firm_aggregates",
            "total_lp":   fa.get("total_lp_investors"),
            "us_lp":      fa.get("us_lp_investors") or 0,
            "non_us_lp":  fa.get("non_us_lp_investors") or 0,
            "no_country": fa.get("lp_investors_no_country_on_file") or 0,
            "total_nav":  fa.get("total_lp_nav") or 0,
            "non_us_nav": fa.get("non_us_lp_nav") or 0,
        }
    return {
        "source": "per_fund_sum",
        "total_lp":   None,
        "us_lp":      sum(d.get("us_lp_investors",              0) or 0 for d in demos.values()),
        "non_us_lp":  sum(d.get("non_us_lp_investors",          0) or 0 for d in demos.values()),
        "no_country": sum(d.get("lp_investors_no_country_on_file", 0) or 0 for d in demos.values()),
        "non_us_nav": sum(d.get("non_us_lp_nav",                0) or 0 for d in demos.values()),
        "total_nav":  sum(d.get("total_lp_nav",                 0) or 0 for d in demos.values()),
    }


def build_firm_overview(wb, fr, demos, firm_aggregates):
    ws = wb.create_sheet("Firm Overview")
    hdr(ws, 1, [""], col_widths=[42, 22, 18, 38])

    fund_count = fr.get("_fund_count", "—")
    row = 1

    lp_agg = _firm_lp_aggregates(firm_aggregates, demos)
    if lp_agg["source"] == "per_fund_sum":
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1,
            value="⚠ Firm-level LP counts may be inflated — firm_aggregates block missing from data file. "
                  "Counts are summed across funds, so LPs committed to multiple funds are double-counted.")
        c.fill = PatternFill("solid", fgColor="FEF3C7")
        c.font = Font(bold=True, color="78350F")
        row += 2

    # Item 5.F
    sect(ws, row, "Item 5.F — Regulatory AUM", 4); row += 1
    hdr(ws, row, ["Item", "Amount", "# Accounts", "Notes"]); row += 1
    for label, amt, accts in [
        ("Discretionary regulatory AUM",  fmt_currency(fr.get("regulatory_aum")), fund_count),
        ("Non-discretionary AUM",          "$0",                                   "0"),
        ("Total regulatory AUM",           fmt_currency(fr.get("regulatory_aum")), fund_count),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        carta(ws, row, 2, amt); carta(ws, row, 3, accts); row += 1

    row += 1
    # Item 5.D
    sect(ws, row, "Item 5.D — Client Types", 4); row += 1
    kv(ws, row, "Client type",         "Pooled Investment Vehicles");  row += 1
    kv(ws, row, "Number of clients",   f"{fund_count} private fund(s)"); row += 1
    if lp_agg.get("total_lp") is not None:
        kv(ws, row, "Total LP beneficial owners (distinct)", fmt_val(lp_agg["total_lp"])); row += 1

    row += 1
    # Item 5.H
    sect(ws, row, "Item 5.H — Non-US Client AUM", 4); row += 1
    hdr(ws, row, ["", "LP Count", "% of LP Investors", "Approx. % of LP NAV"]); row += 1

    us_lp      = lp_agg["us_lp"]
    non_us_lp  = lp_agg["non_us_lp"]
    no_country = lp_agg["no_country"]
    non_us_nav = lp_agg["non_us_nav"]
    total_nav  = lp_agg["total_nav"]
    known      = us_lp + non_us_lp

    for label, count, pct_c, pct_n in [
        ("US persons",          us_lp,      fmt_pct(us_lp / known * 100)      if known else "—", "—"),
        ("Non-US persons",      non_us_lp,  fmt_pct(non_us_lp / known * 100)  if known else "—",
                                            fmt_pct(non_us_nav / total_nav * 100) if total_nav else "—"),
        ("Country not on file", no_country, "—", "—"),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD
        carta(ws, row, 2, count); carta(ws, row, 3, pct_c); carta(ws, row, 4, pct_n); row += 1

    ws.cell(row=row + 1, column=1,
        value="⚠ Non-US % is approximate — based on partner_country field in Carta.").font = ITALIC


def build_per_fund_detail(wb, funds):
    ws = wb.create_sheet("Per-Fund Detail")
    hdr(ws, 1, [""], col_widths=[42, 22, 22, 22, 22])

    row = 1
    for f in funds:
        # Fund banner
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=f"▸ {f.get('fund_name', '—')}")
        c.fill = FUND_FILL; c.font = Font(bold=True, color="FFFFFF", size=12); row += 1

        # Identification
        sect(ws, row, "Identification (Schedule D §7.B.(1))", 5); row += 1
        for label, val, is_carta in [
            ("Legal fund name (verify vs. Carta display name)",  f.get("fund_name"),                  True),
            ("Private Fund ID Number",                           None,                                 False),
            ("Fund type classification",                         f.get("fund_type_classification"),    True),
            ("Investment strategy code (Carta)",                 f.get("investment_strategy_code"),    True),
            ("Legal structure",                                  f.get("legal_structure"),              True),
            ("Formation date",                                   fmt_val(f.get("formation_date")),     True),
            ("Fiscal year end month",                            None,                                 False),
            ("Open to new investors? (Y/N)",                     None,                                 False),
            ("3(c)(1) or 3(c)(7) exemption?",                   None,                                 False),
            ("Form D file number (021- from SEC EDGAR)",         None,                                 False),
            ("Minimum investment amount",                        None,                                 False),
        ]:
            kv(ws, row, label, fmt_val(val) if val else None, is_carta=is_carta); row += 1

        # Balance sheet
        sect(ws, row, "Balance Sheet & Regulatory AUM", 5); row += 1
        hdr(ws, row, ["Field", "Amount", "", "", ""]); row += 1
        for label, key in [
            ("Fair market value (cost + unrealized G/L)",  "fair_market_value"),
            ("Cash",                                        "cash"),
            ("Other assets",                                "other_assets"),
            ("Total gross assets",                          "total_gross_assets"),
            ("Total borrowings outstanding",                "total_borrowings_outstanding"),
            ("Unfunded commitments",                        "unfunded_commitments"),
            ("Regulatory AUM (gross assets + unfunded)",   "regulatory_aum"),
            ("Net asset value (NAV)",                       "net_asset_value"),
            ("LP NAV",                                      "lp_nav"),
            ("GP NAV",                                      "gp_nav"),
        ]:
            ws.cell(row=row, column=1, value=label).font = BOLD
            carta(ws, row, 2, fmt_currency(f.get(key))); row += 1

        # Investor counts
        sect(ws, row, "Beneficial Owners (Q.13)", 5); row += 1
        for label, key in [
            ("LP investors",          "beneficial_owners_lp"),
            ("GP investors",          "beneficial_owners_gp"),
            ("Total beneficial owners", "total_beneficial_owners"),
        ]:
            ws.cell(row=row, column=1, value=label).font = BOLD
            carta(ws, row, 2, fmt_val(f.get(key))); row += 1
        pit = f.get("investor_count_is_point_in_time")
        ws.cell(row=row, column=1,
            value=f"Count basis: {'Point-in-time as of reporting date' if pit else '⚠ Current snapshot — verify against subscription register'}").font = ITALIC
        row += 1

        # Capital activity — annual
        sect(ws, row, "Capital Activity — Annual", 5); row += 1
        hdr(ws, row, ["", "Total", "LP", "GP", ""]); row += 1
        for label, tk, lk, gk in [
            ("Subscriptions",  "annual_subscriptions",  "annual_lp_subscriptions",  "annual_gp_subscriptions"),
            ("Distributions",  "annual_distributions",  "annual_lp_distributions",  "annual_gp_distributions"),
        ]:
            ws.cell(row=row, column=1, value=label).font = BOLD
            carta(ws, row, 2, fmt_currency(f.get(tk)))
            carta(ws, row, 3, fmt_currency(f.get(lk)))
            carta(ws, row, 4, fmt_currency(f.get(gk))); row += 1

        # Capital activity — ITD
        sect(ws, row, "Capital Activity — Inception-to-Date", 5); row += 1
        hdr(ws, row, ["", "Total", "LP", "GP", ""]); row += 1
        for label, tk, lk, gk in [
            ("Contributions",     "contributions_since_inception", "lp_contributions_since_inception", "gp_contributions_since_inception"),
            ("Distributions",     "distributions_since_inception", "lp_distributions_since_inception", None),
            ("Committed capital", "total_committed_capital",       None,                               None),
        ]:
            ws.cell(row=row, column=1, value=label).font = BOLD
            carta(ws, row, 2, fmt_currency(f.get(tk)))
            if lk: carta(ws, row, 3, fmt_currency(f.get(lk)))
            if gk: carta(ws, row, 4, fmt_currency(f.get(gk)))
            row += 1

        # Performance multiples (Form PF reference)
        sect(ws, row, "Performance Multiples (Form PF Reference)", 5); row += 1
        for label, key in [
            ("DPI",     "total_dpi"),
            ("TVPI",    "total_tvpi"),
            ("LP TVPI", "lp_tvpi"),
        ]:
            ws.cell(row=row, column=1, value=label).font = BOLD
            carta(ws, row, 2, fmt_multiple(f.get(key))); row += 1

        # Per-fund manual fields
        sect(ws, row, "Manual Fields — Enter in IARD", 5); row += 1
        for label in [
            "Auditor name, city, country, PCAOB registration #",
            "Custodian name and address",
            "Is custodian a related person? (Y/N)",
            "Frequency of asset valuation",
            "Who performs valuation (internal / third-party)?",
            "Side pocket arrangement? (Y/N)",
            "Gate on investor redemptions? (Y/N)",
            "Fund-of-funds? (Y/N)",
            "Primary adviser for this fund? (Y/N)",
        ]:
            ws.cell(row=row, column=1, value=label).font = BOLD
            manual(ws, row, 2); row += 1

        row += 2  # spacer between funds


def build_investor_demographics(wb, funds, demos):
    ws = wb.create_sheet("Investor Demographics")
    hdr(ws, 1, [
        "Fund", "US LPs", "Non-US LPs", "No Country", "% Non-US (count)",
        "% Non-US (NAV)", "Individuals", "Corporate/LLC", "Trust/Foundation",
        "Pension/Retirement", "Fund Investors",
    ], col_widths=[36, 12, 12, 12, 16, 16, 14, 16, 18, 18, 14])

    sect_row = 1
    ws.insert_rows(1)
    sect(ws, 1, "Schedule D §7.B.(1) — Questions 14–16: Investor Demographics (per fund)", 11)

    row = 3
    for f in funds:
        fund_uuid = f.get("fund_uuid")
        d = demos.get(fund_uuid, {}) if fund_uuid else {}
        ws.cell(row=row, column=1, value=f.get("fund_name", "—")).font = BOLD
        for col, key in enumerate([
            "us_lp_investors", "non_us_lp_investors", "lp_investors_no_country_on_file",
            "pct_non_us_lp_investors", "pct_non_us_lp_nav",
            "individual_investors", "corporate_investors", "trust_foundation_investors",
            "pension_plan_investors", "fund_investors",
        ], 2):
            val = d.get(key)
            display = fmt_pct(val) if key.startswith("pct_") else fmt_val(val)
            carta(ws, row, col, display)
        row += 1

    ws.cell(row=row + 1, column=1,
        value="⚠ Country detection is approximate. Partners with no country on file are excluded from the US/non-US %.").font = ITALIC


def build_asset_composition(wb, fr):
    ws = wb.create_sheet("Asset Composition")
    hdr(ws, 1, [""], col_widths=[50, 22, 20])

    sect(ws, 1, "Schedule D §5.K.(1) — SMA Asset Category Breakdown (Firm Rollup)", 3)
    ws.cell(row=2, column=1,
        value="Note: §5.K.(1) applies to separately managed accounts. If all AUM is in private funds, this section is N/A — confirm with adviser.").font = ITALIC

    hdr(ws, 3, ["Asset Category", "FMV", "% of Total Active FMV"])
    total_fmv = fr.get("total_active_fmv") or 0
    row = 4
    for label, key in [
        ("Exchange-Traded Equity (U.S. and non-U.S.)",           "fmv_exchange_traded_equity"),
        ("Private Equity (non-public ownership interests)",       "fmv_private_equity"),
        ("Securities issued by pooled investment vehicles",       "fmv_pooled_investment_vehicles"),
        ("Options and warrants",                                  "fmv_options_and_warrants"),
        ("Digital assets / cryptocurrency",                       "fmv_digital_assets"),
        ("Other alternatives",                                    "fmv_other_alternatives"),
    ]:
        val = fr.get(key) or 0
        pct = fmt_pct(val / total_fmv * 100) if total_fmv else "—"
        ws.cell(row=row, column=1, value=label).font = BOLD
        carta(ws, row, 2, fmt_currency(val))
        carta(ws, row, 3, pct); row += 1

    ws.cell(row=row, column=1, value="Total Active Investment FMV").font = Font(bold=True)
    carta(ws, row, 2, fmt_currency(total_fmv))
    carta(ws, row, 3, "100%")


def build_manual_fields(wb):
    ws = wb.create_sheet("Manual Fields")
    hdr(ws, 1, ["ADV Item", "Field", "Where to Find It"], col_widths=[12, 48, 42])

    sect(ws, 1, "Fields That Must Be Entered Manually in IARD", 3)
    hdr(ws, 2, ["ADV Item", "Field", "Where to Find It"])

    row = 3
    for item, field, source in [
        ("5.A",      "Total number of employees (full-time + part-time)",        "Internal HR records"),
        ("5.B",      "Employees performing investment advisory functions",         "Internal HR records"),
        ("5.C",      "Types of compensation arrangements (checkboxes)",           "Advisory agreements / fund docs"),
        ("5.E",      "% of regulatory AUM using performance-based fees",          "Advisory agreements"),
        ("5.G",      "Types of advisory services provided (checkboxes)",          "Advisory agreements"),
        ("5.J",      "Sponsor wrap fee programs? (Y/N)",                          "Internal"),
        ("§7.B.(1)", "Private Fund ID Number",                                    "Generate in IARD portal"),
        ("§7.B.(1)", "Legal fund name (verify vs. Carta display name)",           "Fund formation documents"),
        ("§7.B.(1)", "Primary adviser for this fund? (Y/N)",                     "Advisory agreement"),
        ("§7.B.(1)", "Other advisers / sub-advisers",                             "Fund documents"),
        ("§7.B.(1)", "Fiscal year end month",                                     "Fund documents"),
        ("§7.B.(1)", "Currently open to new investors? (Y/N)",                   "Internal"),
        ("§7.B.(1)", "Minimum investment amount",                                  "Subscription agreements"),
        ("§7.B.(1)", "Auditor name, city, country, PCAOB registration #",        "Audit engagement letter"),
        ("§7.B.(1)", "Frequency of asset valuation",                              "Fund documents / valuation policy"),
        ("§7.B.(1)", "Who performs valuation (internal / third-party)?",          "Valuation policy"),
        ("§7.B.(1)", "Is custodian a related person? (Y/N)",                     "Custody agreement"),
        ("§7.B.(1)", "Fund exemption: 3(c)(1) or 3(c)(7)?",                     "Fund formation documents"),
        ("§7.B.(1)", "Form D file number (021- from SEC EDGAR)",                 "SEC EDGAR"),
        ("§7.B.(1)", "Side pocket arrangement? (Y/N)",                            "Fund documents"),
        ("§7.B.(1)", "Gate on investor redemptions? (Y/N)",                      "Fund documents"),
        ("§7.B.(1)", "Fund-of-funds? (Y/N)",                                     "Fund documents"),
        ("Item 6",   "Other business activities",                                 "Internal"),
        ("Item 7.A", "Affiliated advisers and broker-dealers",                    "Internal legal"),
        ("Item 8",   "Participation or interest in client transactions",          "Compliance records"),
        ("Item 9",   "Custodian name, address, sweep arrangements",               "Custody agreement"),
        ("Item 10",  "Control persons",                                            "Ownership records"),
        ("Item 11",  "Disclosure information (regulatory / criminal history)",    "Compliance / legal"),
        ("Sched A/B","Direct and indirect ownership of the adviser",              "Cap table / ownership records"),
    ]:
        ws.cell(row=row, column=1, value=item)
        c = ws.cell(row=row, column=2, value=field)
        c.fill = MANUAL_FILL; c.font = MANUAL_FONT
        ws.cell(row=row, column=3, value=source).font = ITALIC
        row += 1


# ── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Form ADV Excel filing reference")
    parser.add_argument("--data",  required=True, help="Path to form_adv_data.json")
    parser.add_argument("--out",   required=True, help="Output .xlsx path")
    parser.add_argument("--title", default="Form ADV Filing Reference")
    args = parser.parse_args()

    with open(args.data) as fh:
        data = json.load(fh)

    firm_name        = data.get("firm_name", "Firm")
    reporting_date   = data.get("reporting_date", "")
    funds            = data.get("funds", [])
    demos            = data.get("investor_demographics", {})
    firm_aggregates  = data.get("firm_aggregates", {})
    fr = {
        key: sum(f.get(key) or 0 for f in funds)
        for key in [
            "regulatory_aum", "total_active_fmv",
            "fmv_exchange_traded_equity", "fmv_private_equity",
            "fmv_pooled_investment_vehicles", "fmv_options_and_warrants",
            "fmv_digital_assets", "fmv_other_alternatives",
        ]
    }
    fr["_fund_count"] = len(funds)

    wb = Workbook()
    wb.remove(wb.active)

    build_legend(wb, firm_name, reporting_date)
    build_firm_overview(wb, fr, demos, firm_aggregates)
    build_per_fund_detail(wb, funds)
    build_investor_demographics(wb, funds, demos)
    build_asset_composition(wb, fr)
    build_manual_fields(wb)

    out_path = Path(args.out)
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
